import datetime

from openpyxl.reader.excel import load_workbook

from settings.celery import app
from studying_process.models import DirectionOfTraining


class ExcelFile:
    def __init__(self):
        self.filename = f'{datetime.datetime.now()}'
        self.work_book = load_workbook('sample.xlsx')
        self.ws = self.work_book.active

    def __str__(self) -> str:
        return self.filename

    def add_groups(self) -> 'ExcelFile':
        """ Добавить информафию о группах в файл """
        return self

    def add_direction_of_training(self) -> 'ExcelFile':
        """ Добавить информафию о направлениях в файл """
        directions = DirectionOfTraining.objects.select_related('curator').prefetch_related('disciplines').all()
        index_str = 2
        for count_direction, direction in enumerate(directions, start=1):
            self.ws[f'A{index_str}'] = f'{count_direction}.{direction.title}'
            self.ws[f'B{index_str}'] = direction.curator.get_full_name() or direction.curator.username
            index_str += 1
            for count_discipline, discipline in enumerate(direction.disciplines.all(), start=1):
                self.ws[f'A{index_str}'] = f'   {count_discipline}.{discipline.title}'
                index_str += 1
        return self

    def save_file(self):
        """ Сохранить файл """
        self.work_book.save(f"/app/{self.filename}.xlsx")


@app.task
def create_report():
    """ Создание .xlsx отчета администратора """
    excel = ExcelFile()
    excel.add_direction_of_training().add_groups().save_file()
    return 1
